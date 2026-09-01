# %% [markdown]
# # Nature Communications editable tables and source-data workbooks
# Builds release-safe aggregate Excel files from frozen CSV outputs.

# %%
from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NC = Path(os.environ.get('NC_ROOT', Path(__file__).resolve().parents[1])).resolve()
source_root_value = os.environ.get('AKI_SOURCE_ROOT')
if not source_root_value:
    raise RuntimeError('AKI_SOURCE_ROOT must identify the frozen analysis workspace')
ROOT = Path(source_root_value).resolve()
EXTENSION_SOURCE = Path(os.environ.get('EXTENSION_TABLE_ROOT', NC.parent / 'work' / 'tables')).resolve()
release_table_value = os.environ.get('RELEASE_TABLE_ROOT')
RELEASE_TABLE_SOURCE = Path(release_table_value).resolve() if release_table_value else None
MAIN_SOURCE = ROOT / 'delivery' / 'main' / 'tables'
SUPP_SOURCE = ROOT / 'delivery' / 'supplement' / 'tables'
MAIN_OUT = NC / 'tables' / 'main'
SUPP_OUT = NC / 'tables' / 'supplement'
SOURCE_OUT = NC / 'source_data'
for directory in (MAIN_OUT, SUPP_OUT, SOURCE_OUT):
    directory.mkdir(parents=True, exist_ok=True)
(NC / 'qa').mkdir(parents=True, exist_ok=True)

MAIN_TABLES = [
    'Table1_source_model_results.csv',
    'Table2_mixed_MNAR_results.csv',
    'Table3_apparent_vs_reference_recalibration.csv',
    'Table4_reference_sample_design.csv',
]

CROSSDB_SOURCE = ROOT / 'cross_database_external_validation' / 'tables'
CROSSDB_SUPP_TABLES = [
    'Table_inspire_surgical_icu_reference_flow.csv',
    'Table_inspire_surgical_icu_predictor_availability.csv',
    'Table_inspire_locked_external_validation.csv',
    'Table_inspire_locked_external_calibration_curve.csv',
    'Table_inspire_locked_external_decision_curve.csv',
    'Table_inspire_locked_predictor_availability.csv',
    'Table_public_model_to_source_clinical_bridge.csv',
    'Table_public_model_to_source_calibration_curve.csv',
    'Table_public_model_to_source_decision_curve.csv',
    'Table_public_model_to_source_by_centre.csv',
]
EXTENSION_SUPP_TABLES = [
    'Table_method_identification_conditions_v2.csv',
    'Table_primary_selection_reconstruction_decomposition.csv',
    'Table_source_exclusion_linkage_audit.csv',
    'Table_round1_discrepancy_resolution.csv',
    'Table_source_screened_cohort_sensitivity.csv',
    'Table_empirical_schedule_transport.csv',
    'Table_optimized_reference_sampling.csv',
    'Table_source_postdischarge_sensitivity_bounds.csv',
    'Table_hemoglobin_endpoint_replication.csv',
    'Table_hemoglobin_endpoint_flow.csv',
]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b''):
            h.update(block)
    return h.hexdigest()


def safe_sheet(raw: str, used: set[str]) -> str:
    name = re.sub(r'[^A-Za-z0-9_]+', '_', raw).strip('_')[:31] or 'Sheet'
    base, i = name, 1
    while name in used:
        suffix = f'_{i}'
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        if ws.max_row >= 2 and ws.max_column >= 1 and ws.title != 'README':
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(name='Arial', size=10, bold=True, color='000000')
            cell.fill = PatternFill(fill_type='solid', fgColor='E7E7E7')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Arial', size=9, color='000000')
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for idx, column in enumerate(ws.columns, start=1):
            values = [str(c.value) if c.value is not None else '' for c in list(column)[:200]]
            width = min(45, max(10, max((len(v) for v in values), default=10) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width
    if 'INDEX' in wb.sheetnames:
        index = wb['INDEX']
        headers = {cell.value: cell.column for cell in index[1]}
        sheet_column = headers.get('sheet')
        if sheet_column:
            for row in range(2, index.max_row + 1):
                cell = index.cell(row=row, column=sheet_column)
                if cell.value in wb.sheetnames:
                    cell.hyperlink = f"#'{cell.value}'!A1"
                    cell.style = 'Hyperlink'
    wb.save(path)


def workbook_from_csvs(
    csvs: list[Path], output: Path, readme_rows: list[list[str]], *, group_by_display: bool = False
) -> dict:
    used: set[str] = set()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(readme_rows[1:], columns=readme_rows[0]).to_excel(writer, sheet_name='README', index=False)
        used.add('README')
        manifest = []
        if group_by_display:
            grouped: dict[str, list[Path]] = {}
            for csv in csvs:
                display = csv.parent.name
                if not re.match(r'^(?:Supplementary)?Figure\d+$', display):
                    raise RuntimeError(f'Figure source file lacks final-display folder: {csv}')
                grouped.setdefault(display, []).append(csv)
            def display_number(value: str) -> tuple[int, int]:
                match = re.match(r'^(Supplementary)?Figure(\d+)$', value)
                return (1 if match.group(1) else 0, int(match.group(2)))
            for display in sorted(grouped, key=display_number):
                sheet = safe_sheet(display, used)
                parts = []
                for csv in sorted(grouped[display]):
                    frame = pd.read_csv(csv)
                    frame.insert(0, 'panel_source', csv.stem)
                    parts.append(frame)
                    manifest.append({
                        'final_display': display,
                        'source_file': csv.name,
                        'sheet': sheet,
                        'rows': len(frame),
                        'columns': len(frame.columns) - 1,
                        'sha256': sha256(csv),
                    })
                pd.concat(parts, ignore_index=True, sort=False).to_excel(writer, sheet_name=sheet, index=False)
        else:
            for csv in csvs:
                frame = pd.read_csv(csv)
                if csv.name in {
                    'Supplementary_Table_1_method_identification_conditions.csv',
                    'Table_method_identification_conditions_v2.csv',
                }:
                    display = 'Supplementary Table 1'
                    sheet_base = 'Supplementary Table 1'
                elif csv.name == 'Table_primary_selection_reconstruction_decomposition.csv':
                    display = 'Supplementary Table 2'
                    sheet_base = 'Supplementary Table 2'
                else:
                    display = csv.parent.name if re.match(r'^(?:Supplementary)?Figure\d+$', csv.parent.name) else ''
                    stem = re.sub(r'^(?:Supplementary)?Figure\d+[a-z]?(?:-|_)?', '', csv.stem)
                    sheet_base = f'{display}_{stem}' if display else csv.stem
                sheet = safe_sheet(sheet_base, used)
                frame.to_excel(writer, sheet_name=sheet, index=False)
                manifest.append({
                    'final_display': display,
                    'source_file': csv.name,
                    'sheet': sheet,
                    'rows': len(frame),
                    'columns': len(frame.columns),
                    'sha256': sha256(csv),
                })
        pd.DataFrame(manifest).to_excel(writer, sheet_name='INDEX', index=False)
    style_workbook(output)
    return {'output': str(output), 'sha256': sha256(output), 'sheets': manifest}


# %% Copy frozen aggregate tables
main_csvs = []
for name in MAIN_TABLES:
    source = MAIN_SOURCE / name
    if not source.exists():
        raise FileNotFoundError(source)
    target = MAIN_OUT / name
    shutil.copy2(source, target)
    main_csvs.append(target)

compatibility = NC / 'tables' / 'Supplementary_Table_1_method_identification_conditions.csv'
if compatibility.exists():
    target = SUPP_OUT / compatibility.name
    shutil.copy2(compatibility, target)

supp_csvs = []
for source in sorted(SUPP_SOURCE.glob('*.csv')):
    target = SUPP_OUT / source.name
    shutil.copy2(source, target)
    supp_csvs.append(target)
if compatibility.exists():
    supp_csvs.insert(0, SUPP_OUT / compatibility.name)
for name in CROSSDB_SUPP_TABLES:
    source = CROSSDB_SOURCE / name
    if not source.exists():
        raise FileNotFoundError(source)
    target = SUPP_OUT / name
    shutil.copy2(source, target)
    supp_csvs.append(target)
for name in EXTENSION_SUPP_TABLES:
    source = EXTENSION_SOURCE / name
    if not source.exists():
        raise FileNotFoundError(source)
    target = SUPP_OUT / name
    shutil.copy2(source, target)
    supp_csvs.append(target)

# The tagged release is the final authority for any aggregate table that it
# carries. This overlay prevents an older delivery snapshot from silently
# repopulating a submission workbook after a frozen cohort correction.
if RELEASE_TABLE_SOURCE is not None:
    if not RELEASE_TABLE_SOURCE.is_dir():
        raise FileNotFoundError(RELEASE_TABLE_SOURCE)
    supplementary_names = {path.name for path in supp_csvs}
    for source in sorted(RELEASE_TABLE_SOURCE.glob('*.csv')):
        if source.name in MAIN_TABLES:
            shutil.copy2(source, MAIN_OUT / source.name)
        elif source.name in supplementary_names:
            shutil.copy2(source, SUPP_OUT / source.name)
supp_csvs = list(dict.fromkeys(path.resolve() for path in supp_csvs))
# The identification table is Supplementary Table 1 and must be the first data
# sheet rather than an opaque item near the end of a large workbook.
identification = [p for p in supp_csvs if p.name == 'Table_method_identification_conditions_v2.csv']
decomposition = [p for p in supp_csvs if p.name == 'Table_primary_selection_reconstruction_decomposition.csv']
supp_csvs = identification + decomposition + [
    p for p in supp_csvs
    if p.name not in {'Table_method_identification_conditions_v2.csv', 'Table_primary_selection_reconstruction_decomposition.csv'}
]

# %% Main and supplementary table workbooks
main_audit = workbook_from_csvs(
    main_csvs,
    NC / 'tables' / 'Main_Tables.xlsx',
    [['Field', 'Value'], ['Scope', 'Four editable main tables'], ['Data level', 'Aggregate only; no patient-level records'], ['Formatting', 'Black text, grey header, editable cells']],
)

supp_audit = workbook_from_csvs(
    supp_csvs,
    NC / 'tables' / 'Supplementary_Tables.xlsx',
    [
        ['Field', 'Value'],
        ['Scope', 'Complete aggregate supplementary table set'],
        ['Data level', 'Aggregate only; no patient-level records'],
        ['Supplementary Table 1', 'Methods mapped to estimands and identification conditions'],
        ['Supplementary Table 2', 'Common-denominator selection/reconstruction decomposition with fixed-cohort and nested uncertainty'],
    ],
)

# %% Nature Source Data workbook from all submitted panel CSVs
figure_csvs = sorted((NC / 'figures').glob('Figure*/**/*_source_data.csv'))
figure_csvs += sorted((NC / 'figures').glob('SupplementaryFigure*/**/*_source_data.csv'))
# Glob patterns above can overlap; preserve one copy per resolved path.
figure_csvs = list(dict.fromkeys(p.resolve() for p in figure_csvs))
source_audit = workbook_from_csvs(
    figure_csvs,
    SOURCE_OUT / 'Source_Data.xlsx',
    [
        ['Field', 'Value'],
        ['Article title', 'Transported outcome-measurement schedules can alter calibration of clinical prediction models'],
        ['Scope', 'Aggregate source values for every main and supplementary figure panel; INDEX maps final display numbers to sheets'],
        ['Data level', 'Aggregate or Monte Carlo summary only; no stable patient identifiers or row-level predictions'],
        ['Endpoint note', 'Public endpoints are operational creatinine references, not biological truth or expert-adjudicated full KDIGO'],
    ],
    group_by_display=True,
)

audit = {
    'status': 'PASS',
    'main_tables': main_audit,
    'supplementary_tables': supp_audit,
    'source_data': source_audit,
    'main_csv_count': len(main_csvs),
    'supplementary_csv_count': len(supp_csvs),
    'figure_source_csv_count': len(figure_csvs),
}
(NC / 'qa' / 'WORKBOOK_BUILD_AUDIT.json').write_text(json.dumps(audit, indent=2) + '\n')
print(json.dumps({k: audit[k] for k in ['status', 'main_csv_count', 'supplementary_csv_count', 'figure_source_csv_count']}, indent=2))
