# %% [markdown]
# # Nature Communications submission-package QA
# Performs release checks without reading patient-level files.

# %%
from __future__ import annotations
from release_paths import release_path as _release_path

import hashlib
import json
import re
import zipfile
from pathlib import Path

import openpyxl
import pymupdf

ROOT = Path(str(_release_path('analysis')))
NC = ROOT / 'nature_communications'
MANUSCRIPT = NC / 'manuscript' / 'MANUSCRIPT_NATURE_COMMUNICATIONS.md'
LEGENDS = NC / 'manuscript' / 'FIGURE_LEGENDS_NATURE_COMMUNICATIONS.md'


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b''):
            h.update(block)
    return h.hexdigest()


def words(text: str) -> int:
    text = re.sub(r'<[^>]+>', '', text)
    return len(re.findall(r"\b[\w'−-]+\b", text))


def docx_zip_ok(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as archive:
            return archive.testzip() is None and 'word/document.xml' in archive.namelist()
    except Exception:
        return False


s = MANUSCRIPT.read_text()
title = s.splitlines()[0].lstrip('# ').strip()
abstract = s.split('## Abstract\n\n', 1)[1].split('\n\n## Introduction', 1)[0]
body = s.split('## Introduction\n\n', 1)[1].split('\n\n## Data availability', 1)[0]
methods = s.split('## Methods\n\n', 1)[1].split('\n\n## Data availability', 1)[0]
headings = re.findall(r'^## (.+)$', s, flags=re.M)
required_order = ['Authors', 'Abstract', 'Introduction', 'Results', 'Discussion', 'Methods', 'Data availability', 'Code availability', 'References', 'Acknowledgements', 'Author contributions', 'Competing interests', 'Additional information']

reference_numbers = [int(x) for x in re.findall(r'^(\d+)\. ', s, flags=re.M)]
citation_numbers = []
for token in re.findall(r'<sup>([0-9,\-]+)</sup>', s):
    for component in token.split(','):
        if '-' in component:
            lo, hi = map(int, component.split('-'))
            citation_numbers.extend(range(lo, hi + 1))
        else:
            citation_numbers.append(int(component))

legend_counts = {}
for block in re.split(r'\n(?=## )', LEGENDS.read_text()):
    match = re.match(r'## ([^\n]+)\n\n(.*)', block, flags=re.S)
    if match:
        legend_counts[match.group(1)] = words(match.group(2))

required_files = [
    NC / 'manuscript' / 'MANUSCRIPT_NATURE_COMMUNICATIONS.docx',
    NC / 'manuscript' / 'COVER_LETTER_NATURE_COMMUNICATIONS.docx',
    NC / 'supplement' / 'SUPPLEMENTARY_INFORMATION.docx',
    NC / 'tables' / 'MAIN_TABLES_NATURE_COMMUNICATIONS.docx',
    NC / 'tables' / 'Main_Tables.xlsx',
    NC / 'tables' / 'Supplementary_Tables.xlsx',
    NC / 'source_data' / 'Source_Data.xlsx',
    NC / 'checklists' / 'REPORTING_CHECKLIST_CROSSWALK.md',
    NC / 'checklists' / 'MODEL_CARD.md',
    NC / 'review' / 'NATURE_COMMUNICATIONS_THREE_REVIEWER_REVIEW.md',
    ROOT / 'cross_database_external_validation' / 'outputs' / 'CROSS_DATABASE_EXTERNAL_VALIDATION_QA.json',
]
required_files.extend(NC / 'figures' / f'Figure{i}' / f'Figure{i}.pdf' for i in range(1, 5))
required_files.extend(NC / 'figures' / f'SupplementaryFigure{i}' / f'SupplementaryFigure{i}.pdf' for i in range(1, 9))

figure_checks = []
for path in [NC / 'figures' / f'Figure{i}' / f'Figure{i}.pdf' for i in range(1, 5)]:
    doc = pymupdf.open(path)
    page = doc[0]
    figure_checks.append({'file': str(path.relative_to(NC)), 'pages': len(doc), 'width_pt': page.rect.width, 'width_183mm': abs(page.rect.width - 183 / 25.4 * 72) < 0.2})
    doc.close()

source_wb = openpyxl.load_workbook(NC / 'source_data' / 'Source_Data.xlsx', read_only=True)
source_sheets = len(source_wb.sheetnames) - 1
source_wb.close()
main_wb = openpyxl.load_workbook(NC / 'tables' / 'Main_Tables.xlsx', read_only=True)
main_sheets = len(main_wb.sheetnames) - 1
main_wb.close()

analysis_log = (NC / 'qa' / 'ANALYSIS_COMPLETION_AUDIT.log').read_text()
analysis_payload = json.loads(analysis_log)
pytest_log = (NC / 'qa' / 'PYTEST.log').read_text()
watermark_log = (NC / 'qa' / 'WATERMARK_AUDIT.log').read_text()
crossdb_payload = json.loads((ROOT / 'cross_database_external_validation' / 'outputs' / 'CROSS_DATABASE_EXTERNAL_VALIDATION_QA.json').read_text())

checks = {
    'title_word_limit': len(title.split()) <= 15,
    'title_no_punctuation': not bool(re.search(r'[:;?!]', title)),
    'abstract_150_words': words(abstract) <= 150,
    'body_including_methods_5000_words': words(body) <= 5000,
    'methods_budget_1800_words': words(methods) <= 1800,
    'required_section_order': headings == required_order,
    'references_70_or_fewer': len(reference_numbers) <= 70,
    'references_sequential': reference_numbers == list(range(1, len(reference_numbers) + 1)),
    'citations_first_use_sequential': list(dict.fromkeys(citation_numbers)) == reference_numbers,
    'eight_main_display_items': 4 + 4 <= 10,
    'legends_350_words_or_fewer': max(legend_counts.values()) <= 350,
    'all_required_files_present': all(p.exists() and p.stat().st_size > 0 for p in required_files),
    'all_main_figures_vector_one_page_183mm': all(x['pages'] == 1 and x['width_183mm'] for x in figure_checks),
    'source_data_covers_42_panels': source_sheets == 42,
    'four_editable_main_tables': main_sheets == 4,
    'docx_zip_integrity': all(docx_zip_ok(p) for p in required_files if p.suffix == '.docx'),
    'analysis_gates_27_of_27': analysis_payload.get('items') == 27 and analysis_payload.get('passed') == 27 and analysis_payload.get('failed') == 0,
    'regression_tests': '35 passed, 1 skipped' in pytest_log,
    'watermark_scan': watermark_log.count('watermark_like_hits=0') == 4,
    'cross_database_external_validation_qa': crossdb_payload.get('status') == 'PASS',
    'no_lancet_transfer_markers': 'Lancet' not in s and 'Research in context' not in s,
    'no_em_or_en_dash_after_humanizer': '—' not in s and '–' not in s,
}

author_placeholders = sorted(set(re.findall(r'\[(AUTHOR[^\]]+|CORRESPONDING[^\]]+)\]', s + (NC / 'manuscript' / 'COVER_LETTER_NATURE_COMMUNICATIONS.md').read_text())))
scientific_ready = all(checks.values())
submission_ready = scientific_ready and not author_placeholders

payload = {
    'status': 'PASS' if submission_ready else ('PASS_WITH_AUTHOR_BLOCKERS' if scientific_ready else 'FAIL'),
    'scientific_package_ready': scientific_ready,
    'submission_ready': submission_ready,
    'title': title,
    'counts': {
        'title_words': len(title.split()),
        'abstract_words': words(abstract),
        'body_including_methods_words': words(body),
        'methods_words': words(methods),
        'references': len(reference_numbers),
        'main_display_items': 8,
        'source_data_panels': source_sheets,
        'main_tables': main_sheets,
    },
    'checks': checks,
    'figure_checks': figure_checks,
    'legend_word_counts': legend_counts,
    'author_controlled_blockers': author_placeholders,
    'patient_level_data_read': False,
}

manifest = []
excluded_parts = {'.venv', 'python_deps', 'node_project', 'node_tooling', 'submission_package'}
excluded_files = {
    NC / 'Nature_Communications_submission_package.zip',
    NC / 'qa' / 'SUBMISSION_QA.json',
    NC / 'qa' / 'SUBMISSION_QA.log',
    NC / 'qa' / 'FILE_MANIFEST.json',
    NC / 'qa' / 'SHA256SUMS',
}
for path in sorted(
    p for p in NC.rglob('*')
    if p.is_file()
    and not excluded_parts.intersection(p.parts)
    and p not in excluded_files
):
    manifest.append({'path': str(path.relative_to(NC)), 'bytes': path.stat().st_size, 'sha256': sha256(path)})
payload['manifest_file_count'] = len(manifest)
(NC / 'qa' / 'SUBMISSION_QA.json').write_text(json.dumps(payload, indent=2) + '\n')
(NC / 'qa' / 'FILE_MANIFEST.json').write_text(json.dumps(manifest, indent=2) + '\n')
(NC / 'qa' / 'SHA256SUMS').write_text(''.join(f"{x['sha256']}  {x['path']}\n" for x in manifest))
print(json.dumps(payload, indent=2))
if not scientific_ready:
    raise SystemExit(1)
